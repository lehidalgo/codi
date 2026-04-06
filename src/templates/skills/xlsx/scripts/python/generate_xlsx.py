#!/usr/bin/env python3
"""
generate_xlsx.py — Brand+theme-aware XLSX generator.
Usage: python3 generate_xlsx.py --content content.json [--tokens brand_tokens.json] [--theme dark|light] --output out.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def to_hex8(hex_color: str) -> str:
    """Convert #RRGGBB to FFRRGGBB for openpyxl."""
    return "FF" + hex_color.lstrip("#").upper()


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--content", required=True)
    p.add_argument("--tokens")
    p.add_argument("--theme", default="dark", choices=["dark", "light"])
    p.add_argument("--output", required=True)
    return p.parse_args()


def main():
    args = parse_args()
    default_tokens = Path(__file__).parent.parent / "brand_tokens.json"
    tokens_path = Path(args.tokens) if args.tokens else default_tokens
    tokens = json.loads(tokens_path.read_text())
    T = tokens["themes"][args.theme]
    F = tokens["fonts"]
    content = json.loads(Path(args.content).read_text())

    wb = Workbook()
    wb.remove(wb.active)

    for sec in content["sections"]:
        sheet_name = (sec.get("heading") or sec.get("label") or "Sheet")[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = T["accent"].lstrip("#")

        header_fill = PatternFill("solid", fgColor=to_hex8(T["primary"]))
        header_font = Font(
            bold=True,
            color=to_hex8(T["text_primary"]),
            name=F["fallback_sans"],
            size=11,
        )
        accent_border = Border(bottom=Side(style="thin", color=T["accent"].lstrip("#")))

        ws.append(["#", "Item", "Details"])
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = accent_border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[ws.max_row].height = 22

        for idx, item in enumerate(sec.get("items", [])):
            fill_hex = T["surface"] if idx % 2 == 1 else T["background"]
            row_fill = PatternFill("solid", fgColor=to_hex8(fill_hex))
            row_font = Font(
                color=to_hex8(T["text_primary"]), name=F["fallback_sans"], size=10
            )
            ws.append([idx + 1, item, sec.get("body", "")])
            for cell in ws[ws.max_row]:
                cell.fill = row_fill
                cell.font = row_font

        if sec.get("callout"):
            ws.append([sec["callout"]])
            note_row = ws.max_row
            ws.merge_cells(f"A{note_row}:C{note_row}")
            ws[f"A{note_row}"].font = Font(
                italic=True,
                color=to_hex8(T["accent"]),
                name=F["fallback_sans"],
                size=10,
            )

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 40

    wb.save(args.output)
    print(f"XLSX written: {args.output} ({len(content['sections'])} sheets)")


if __name__ == "__main__":
    main()
